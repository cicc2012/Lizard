import mimetypes
import os
import time
from pathlib import Path
from urllib.parse import urlparse

import requests
from openpyxl import Workbook, load_workbook

# Customize your search query here
# Follow the iNaturalist API documentation for more options:
# https://api.inaturalist.org/v1/docs/#!/Observations/get_observations
params = {
    'taxon_name': 'Crotaphytus collaris',
    'per_page': 50,  # Number of results to return in a page. Max is 200.
    'page': 1,
    'has[]': 'photos',  # Only return observations that have photos.
}

# TODO: Change these limits as desired
PAGE_LIMIT = 20
COUNT_LIMIT = 70
skip_count = 700  # Skip the first N images in this run
ENABLE_OKLAHOMA_FILTER = False
OKLAHOMA_BBOX = {
    'min_lat': 33.615,
    'max_lat': 37.002,
    'min_lng': -103.002,
    'max_lng': -94.430,
}

# Specify the directory to save images
save_directory = Path(r"C:\Users\szhang10\Documents\Test\Lizard\data\images\26Apr14") / params['taxon_name'].replace(' ', '_')
save_directory.mkdir(parents=True, exist_ok=True)

log_file = save_directory / "image_log.xlsx"

existing_files_dir = Path(r"F:\Test\Lizard\test\dataset\images\train")
if existing_files_dir.exists():
    existing_files = {f.name for f in existing_files_dir.iterdir() if f.is_file()}
else:
    existing_files = set()
    print(f"Warning: existing_files_dir does not exist: {existing_files_dir}")

existing_files.update({f.name for f in save_directory.iterdir() if f.is_file()})

# Excel workbook initialization
if os.path.exists(log_file):
    wb = load_workbook(log_file)
    ws = wb.active
else:
    wb = Workbook()
    ws = wb.active
    ws.title = "Image Log"
    ws.append([
        "Image Filename",
        "Image URL",
        "Observation ID",
        "Photo ID",
        "Time Used (ms)",
        "Location",
        "Place Guess",
        "Observed On",
        "Observer Username",
        "Quality Grade",
        "Common Name",
        "Scientific Name",
    ])
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 22
    ws.column_dimensions['G'].width = 30
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 18
    ws.column_dimensions['J'].width = 15
    ws.column_dimensions['K'].width = 22
    ws.column_dimensions['L'].width = 22

MAX_RETRIES = 3
RETRY_DELAY = 5
REQUEST_TIMEOUT = 20
RATE_LIMIT_SECONDS = 1.1
USER_AGENT = "iNaturalistImageDownloader/1.0 (+https://github.com/)"

session = requests.Session()
session.headers.update({"User-Agent": USER_AGENT})


def safe_request(url, params=None, timeout=REQUEST_TIMEOUT):
    attempts = 0
    while attempts < MAX_RETRIES:
        try:
            response = session.get(url, params=params, timeout=timeout)
            if response.status_code == 429:
                wait = int(response.headers.get("Retry-After", RETRY_DELAY))
                print(f"Rate limited by server, waiting {wait} seconds...")
                time.sleep(wait)
                attempts += 1
                continue
            response.raise_for_status()
            return response
        except requests.RequestException as exc:
            attempts += 1
            print(f"Request failed ({attempts}/{MAX_RETRIES}): {exc}")
            if attempts >= MAX_RETRIES:
                print("Max retries reached for this request.")
                return None
            time.sleep(RETRY_DELAY)
    return None


def is_duplicate_image(filename, existing_names=None):
    existing_names = existing_names or set()
    stem = Path(filename).stem
    if filename in existing_names:
        return True
    for existing_name in existing_names:
        if Path(existing_name).stem == stem:
            return True
    return False


def get_image_extension(response, url):
    content_type = response.headers.get("Content-Type", "").split(";", 1)[0].strip().lower()
    ext_map = {
        "image/jpeg": ".jpg",
        "image/jpg": ".jpg",
        "image/png": ".png",
        "image/webp": ".webp",
        "image/gif": ".gif",
        "image/tiff": ".tiff",
        "image/x-tiff": ".tiff",
    }
    extension = ext_map.get(content_type)
    if not extension and content_type:
        extension = mimetypes.guess_extension(content_type)
    if not extension:
        url_path = urlparse(url).path
        extension = Path(url_path).suffix
    if not extension:
        extension = ".jpg"
    return extension.lower()


def parse_location_coords(obs):
    location = obs.get("private_location") or obs.get("location")
    if not location:
        return None
    parts = [part.strip() for part in str(location).split(",")]
    if len(parts) != 2:
        return None
    try:
        return float(parts[0]), float(parts[1])
    except ValueError:
        return None


def is_oklahoma_observation(obs):
    coords = parse_location_coords(obs)
    if coords:
        lat, lng = coords
        return (
            OKLAHOMA_BBOX['min_lat'] <= lat <= OKLAHOMA_BBOX['max_lat']
            and OKLAHOMA_BBOX['min_lng'] <= lng <= OKLAHOMA_BBOX['max_lng']
        )
    place_guess = str(obs.get("place_guess", "")).lower()
    return "oklahoma" in place_guess


def save_workbook():
    try:
        wb.save(log_file)
    except Exception as exc:
        print(f"Failed to save workbook: {exc}")


downloaded_count = 0
skipped_count = 0
start_time_global = time.time()

while params['page'] <= PAGE_LIMIT and downloaded_count < COUNT_LIMIT:
    print(f"Fetching observations page {params['page']}...")
    response = safe_request("https://api.inaturalist.org/v1/observations", params=params)
    if response is None:
        print("Stopping because observation request failed.")
        break

    try:
        payload = response.json()
    except ValueError as exc:
        print(f"Failed to decode JSON response: {exc}")
        break

    results = payload.get("results", [])
    if not results:
        print("No observations returned on this page.")
        break

    for obs in results:
        if downloaded_count >= COUNT_LIMIT:
            break

        observation_id = obs.get("id", "unknown")
        if ENABLE_OKLAHOMA_FILTER and not is_oklahoma_observation(obs):
            print(f"Skipping observation {observation_id} because it is outside Oklahoma.")
            continue

        location = obs.get("private_location") or obs.get("location") or "N/A"
        place_guess = obs.get("place_guess") or "N/A"
        observed_on = obs.get("observed_on") or obs.get("observed_at") or "N/A"
        observer_username = obs.get("user", {}).get("login", "N/A")
        quality_grade = obs.get("quality_grade", "N/A")
        taxon_common_name = (
            obs.get("taxon", {}).get("preferred_common_name")
            or obs.get("taxon", {}).get("name")
            or "N/A"
        )
        scientific_name = obs.get("taxon", {}).get("scientific_name", "N/A")

        for photo in obs.get("photos", []):
            if downloaded_count >= COUNT_LIMIT:
                break

            photo_id = photo.get("id", "unknown")
            photo_url = photo.get("url")
            if not photo_url:
                continue

            image_url = photo_url.replace("square", "original")
            start_time = time.time()
            image_response = safe_request(image_url)
            if image_response is None:
                print(f"Skipping image {observation_id}_{photo_id} due to download error.")
                continue

            extension = get_image_extension(image_response, image_url)
            base_filename = f"{observation_id}_{photo_id}{extension}"
            if is_duplicate_image(base_filename, existing_files):
                print(f"Image already exists from previous downloads: {base_filename}. Skipping.")
                continue

            img_path = save_directory / base_filename

            if skipped_count < skip_count:
                skipped_count += 1
                print(f"Skipping image {observation_id}_{photo_id}. Skipped count: {skipped_count}")
                continue

            try:
                with open(img_path, "wb") as f:
                    f.write(image_response.content)
            except OSError as exc:
                print(f"Failed to write image file {img_path}: {exc}")
                continue

            end_time = time.time()
            time_used_ms = (end_time - start_time) * 1000

            ws.append([
                str(img_path),
                image_url,
                observation_id,
                photo_id,
                f"{time_used_ms:.2f}",
                location,
                place_guess,
                observed_on,
                observer_username,
                quality_grade,
                taxon_common_name,
                scientific_name,
            ])

            downloaded_count += 1
            existing_files.add(base_filename)
            print(f"Downloaded {img_path} in {time_used_ms:.2f} ms ({downloaded_count}/{COUNT_LIMIT})")

            save_workbook()
            time.sleep(max(0, RATE_LIMIT_SECONDS - time_used_ms / 1000))

    params['page'] += 1

save_workbook()
end_time_global = time.time()
total_time_global = end_time_global - start_time_global
print(f"Downloaded {downloaded_count} images in {total_time_global:.2f} seconds. Log saved to {log_file}.")